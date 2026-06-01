"""
Загрузка и парсинг файлов различных форматов.
"""

import os
from pathlib import Path
from typing import Dict, List, Optional

import docx
from PyPDF2 import PdfReader

from src.core.config import Config


class DocumentLoader:
    """Слой I/O: Отвечает за доступ к файловой системе и парсинг форматов."""

    def __init__(self, config: Config):
        self.config = config
        self.data_path = Path(config.data_path)

    def _load_ragignore(self) -> list[str]:
        """Читает паттерны исключений из data/.ragignore (синтаксис fnmatch)."""
        ragignore_path = self.data_path / ".ragignore"
        if not ragignore_path.exists():
            return []
        lines = ragignore_path.read_text(encoding="utf-8").splitlines()
        return [line.strip() for line in lines if line.strip() and not line.startswith("#")]

    def list_files(self) -> List[Path]:
        """Список всех поддерживаемых файлов в data/, исключая .ragignore паттерны."""
        import fnmatch

        if not self.data_path.exists():
            self.data_path.mkdir(parents=True)
            return []

        supported = {".txt", ".pdf", ".doc", ".docx", ".xlsx", ".md"}
        ignore_patterns = self._load_ragignore()
        files: List[Path] = []

        for root, _, filenames in os.walk(self.data_path):
            for filename in filenames:
                filepath = Path(root) / filename
                if filepath.suffix.lower() not in supported:
                    continue
                rel_path = str(filepath.relative_to(self.data_path))
                if any(fnmatch.fnmatch(rel_path, p) or fnmatch.fnmatch(filename, p) for p in ignore_patterns):
                    continue
                files.append(filepath)

        return files

    def load_documents(self, files: Optional[List[Path]] = None) -> List[Dict]:
        """Загрузка документов с парсингом метаданных."""
        documents = []

        if files is None:
            files = self.list_files()

        for filepath in files:
            text = self._load_file(filepath)
            if not text:
                continue

            rel_source = filepath.relative_to(self.data_path)

            # Парсинг метаданных из Markdown
            metadata = {}
            if filepath.suffix.lower() == ".md" and text.strip().startswith("---"):
                text, metadata = self._parse_markdown_metadata(text, filepath.name)

            documents.append({"text": text, "source": str(rel_source).replace("\\", "/"), "metadata": metadata})

        return documents

    def _load_file(self, path: Path) -> str:
        """Загрузка документа по расширению."""
        ext = path.suffix.lower()
        try:
            if ext in {".txt", ".md"}:
                return path.read_text(encoding="utf-8")
            if ext == ".pdf":
                return self._load_pdf(path)
            if ext == ".doc":
                return self._load_doc(path)
            if ext == ".docx":
                return self._load_docx(path)
            if ext == ".xlsx":
                return self._load_xlsx(path)
        except Exception as e:
            print(f"  ⚠ Ошибка загрузки {path.name}: {e}")
            return ""
        return ""

    def _load_pdf(self, path: Path) -> str:
        text = ""
        pdf = PdfReader(str(path))
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text

    def _load_doc(self, path: Path) -> str:
        try:
            import textract

            return textract.process(str(path)).decode("utf-8")
        except Exception:
            # Fallback или просто пропуск, как было в оригинале
            return ""

    def _load_docx(self, path: Path) -> str:
        doc = docx.Document(str(path))
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    def _load_xlsx(self, path: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), data_only=True)
        text_parts = []
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or not any(rows[0]):
                continue
            header = rows[0]
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                pairs = [f"{h}: {v}" for h, v in zip(header, row, strict=False) if v is not None and str(v).strip()]
                if pairs:
                    text_parts.append(" | ".join(pairs))
        return "\n".join(text_parts)

    def _parse_markdown_metadata(self, text: str, filename: str) -> tuple[str, dict]:
        metadata = {}
        try:
            parts = text.split("---", 2)
            if len(parts) >= 3:
                yaml_content = parts[1]
                text = parts[2].strip()
                for line in yaml_content.strip().split("\n"):
                    if ":" in line:
                        k, v = line.split(":", 1)
                        val = v.strip()
                        # Убираем кавычки, если они есть
                        if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
                            val = val[1:-1].strip()
                        metadata[k.strip()] = val
        except Exception as e:
            print(f"  ⚠ Ошибка парсинга метаданных в {filename}: {e}")
        return text, metadata
